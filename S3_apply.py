import boto3
import logging
import pandas as pd
from datetime import datetime
from botocore.exceptions import ClientError
from time import strftime 
from tqdm import tqdm
import json
import ast
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
logname = datetime.now().strftime('logfile_S3_Apply_%H_%M_%S_%d_%m_%Y.log')
print(logname)
logging.basicConfig(filename=logname,level=logging.INFO,filemode='w', format='%(levelname)s: %(asctime)s: %(message)s' ,force=True)
logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client('s3')
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
iam = boto3.resource('iam')
client = boto3.client('sts')
ignorelist=[]
servicelist=['vpc','s3','elb','CloudTrail','rds','']
region=['us-east-1','us-west-1','ca-central-1','eu-west-2','us-west-2','us-east-2','ap-south-1']
column_list=[None, 'ID', 'Filter', 'Status', 'AbortIncompleteMultipartUpload', 'Prefix', 'Expiration', 'Transitions', 'NoncurrentVersionExpiration', 'NoncurrentVersionTransitions']
stdpname = ['MMS_Versioning_Policy_3v_retains','MMSDeletionStandardPolicy','MMSDeleteMarkers','AbortIncompleteMultipartUploadsRule']
## Global list variable to keep track of the Bucket Name, Transition Days, StorageClass, Status  
TransitionStatus = []

policy = {}
ruledict= {}


MMSMoveCustomPolicy = {
    'Rules': [
        {
        'ID': 'MMSMoveCustomPolicy',
        'Expiration': {'Days': 2555},
        'Filter': {'ObjectSizeGreaterThan': 131072}, 
        'Status': 'Enabled',
        'Transitions': [{'Days': 121, 'StorageClass': 'GLACIER_IR'}]
        }
        ]
    }



def put_bucket_lifecycle_configuration_custom(Name, lifecycle_config,pname):
    ownerAccountId = getAccountID()
    try:
        result = s3.get_bucket_lifecycle_configuration(Bucket=Name, ExpectedBucketOwner=ownerAccountId)
        #print(result)
        Rules= result['Rules']          
        for target in Rules:
                try:
                    #print(target['ID'],pname)               
                    if target['ID'] == pname:
                        print('0-ReCreating LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        logging.info('0-ReCreating LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        Rules.remove(target)
                        s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                except  KeyError:
                    print('Error')
        
        
        #print(result)        
        Rules.append(lifecycle_config)
        #print(Rules)
        lcp = s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
    except ClientError as err:
        #print(err.response)
        if err.response['Error']['Code'] == 'NoSuchLifecycleConfiguration':
            s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules': lifecycle_config['Rules'] })            
        elif err.response['Error']['Code'] == 'InvalidRequest':
            logging.info(err.response)
        elif err.response['Error']['Code'] == 'MalformedXML':
            logging.info( err.response['Error']['Code'] + ' Rule info is  not well-formed or did not validate against our published schema')




# method to determine accountID 
def getAccountID():
    account_id = client.get_caller_identity()
    return(account_id['Account'])
    #account_id = iam.CurrentUser().arn.split(':')[4]
    #return(account_id)



    

def updatelcp(filename):
    wb = load_workbook(filename)
    ws = wb.active

    max_row=ws.max_row
    max_col=ws.max_column

    #print(max_row,max_col)
            
            
    #lifecycle_config={}                
    for row_cells in ws.iter_rows(min_row=2, max_row=max_row):
            counter = 0
            lifecycle_config={} 
            for i, cell in enumerate(row_cells):

                counter += 1
                #print(get_column_letter(counter) + str(1))
                colname = ws[get_column_letter(counter) + str(1)]
                #print(colname.value,cell.value)
                if colname.value is None:
                    bucket_name = cell.value
                    
                    #print(bucket_name)                    
                elif colname.value == 'ID':
                    ID = cell.value
                    lifecycle_config['ID'] = ID
                    #print(ID)
                elif colname.value == 'Expiration':
                    Expiration = cell.value
                    if cell.value is not None:
                        lifecycle_config['Expiration'] = ast.literal_eval(Expiration)
                    else :
                        lifecycle_config['Expiration'] = {}
                elif colname.value == 'Filter':
                    Filter = cell.value
                    if cell.value is not None:
                        lifecycle_config['Filter'] = ast.literal_eval(Filter)
                    else :
                        lifecycle_config['Filter'] = {}
                elif colname.value == 'Status':
                    Status = cell.value
                    lifecycle_config['Status'] = Status
                elif colname.value == 'Transitions':
                    Transitions = cell.value
                    if cell.value is not None:
                        lifecycle_config['Transitions'] = ast.literal_eval(Transitions)
                    else:
                        lifecycle_config['Transitions'] = []
                    #print(Transitions) 
            #lifecycle_config = f"""{{'Rules':[{{'ID': '{ID}','Expiration': {Expiration},'Filter': {Filter},'Status': '{Status}','Transitions': {Transitions}}}]}}"""
            #lifecycle_config = json.loads(f"""{{"Rules":[{{"ID": "{ID}","Expiration": "{Expiration}","Filter": "{Filter}","Status": "{Status}","Transitions": "{Transitions}"}}]}}""")
            #print(lifecycle_config)
    
            put_bucket_lifecycle_configuration_custom(bucket_name, lifecycle_config,ID)

                                                          
                
        
        

    
def main():
    file_name= 'postchange' + '_' + getAccountID() + ".xlsx"
    updatelcp(file_name)
    
    
    
if __name__ == "__main__":
    main()