# This script is intended to simplify the process of adding Intelligent Tiering Lifecycle Policies 
# to S3 buckets for all objects >128KB. See https://aws.amazon.com/s3/storage-classes/intelligent-tiering/ for
# details on Intelligent Tiering for S3 Buckets. 
# The Amazon S3 Intelligent-Tiering storage class is designed to optimize storage costs by automatically moving data to the most cost-effective access tier when access patterns change. 
# For a small monthly object monitoring and automation charge, S3 Intelligent-Tiering monitors access patterns and automatically moves objects that have not been accessed to lower cost access tiers. 
# S3 Intelligent-Tiering is the ideal storage class for data with unknown, changing, or unpredictable access patterns, independent of object size or retention period. 
# You can use S3 Intelligent-Tiering as the default storage class for data lakes, analytics, and new applications.  


import boto3
import pandas as pd
from datetime import datetime
from botocore.exceptions import ClientError
from time import strftime 
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font

s3 = boto3.client('s3')
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
iam = boto3.resource('iam')
client = boto3.client('sts')
ignorelist=[]
servicelist=['vpc','s3','elb','CloudTrail','rds']
region=['us-east-1','us-west-1','ca-central-1','eu-west-2','us-west-2','us-east-2']
column_list=[None, 'ID', 'Filter', 'Status', 'AbortIncompleteMultipartUpload', 'Prefix', 'Expiration', 'Transitions', 'NoncurrentVersionExpiration', 'NoncurrentVersionTransitions']

MMSVersioningPolicy = {
    'Rules': [
        {'ID': 'MMSVersioningPolicy',
         'Expiration': {'Days': 31},
         'Filter': {},
         'Status': 'Enabled', 
         'NoncurrentVersionExpiration': {'NoncurrentDays': 31, 'NewerNoncurrentVersions': 1}
        }
    ]}

MMSMPUPolicy = {
    'Rules': [
        {
          'ID': 'AbortIncompleteMultipartUploadsRule',
          'Expiration': {'ExpiredObjectDeleteMarker': False},
          'Filter': {},
          'Status': 'Enabled',
          'AbortIncompleteMultipartUpload': {'DaysAfterInitiation': 7}           
        }        
        ]
    }

MMSDeleteMarkers= {
    'Rules': [
        {
          'ID': 'MMSDeleteMarkers',
          'Expiration': {'ExpiredObjectDeleteMarker': True},
          'Filter': {},
          'Status': 'Enabled'         
        }        
        ]
    }

MMSDeletionStandardPolicy = {
    'Rules': [
        {
        'ID': 'MMSDeletionStandardPolicy',
        'Expiration': {'Days': 2555},
        'Filter': {}, 
        'Status': 'Enabled'
        }
        ]
    }

MMSMoveCustomPolicy = {
    'Rules': [
        {
        'ID': 'MMSMoveCustomPolicy',
        'Filter': {'ObjectSizeGreaterThan': 131072}, 
        'Status': 'Enabled',
        'Transitions': [{'Days': 120, 'StorageClass': 'GLACIER_IR'}]
        }
        ]
    }

def put_bucket_lifecycle_configuration_standard(Name, lifecycle_config):
    ownerAccountId = getAccountID()
    try:
        result = s3.get_bucket_lifecycle_configuration(Bucket=Name, ExpectedBucketOwner=ownerAccountId)
        #print(result)
        Rules= result['Rules']
        if lifecycle_config['Rules'][0]['ID'] == 'MMSDeleteMarkers':            
            for target in Rules:
                try:
                    #print(target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'])               
                    if target['Expiration']['ExpiredObjectDeleteMarker'] == 'Ture' or target['ID'] not in ['MMSDeletionStandardPolicy','MMSDeleteMarkers','AbortIncompleteMultipartUploadsRule','MMSVersioningPolicy']:
                        print('0-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        Rules.remove(target)
                        s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                except  KeyError:
                    continue
        elif lifecycle_config['Rules'][0]['ID'] == 'AbortIncompleteMultipartUploadsRule':
            for target in Rules:
                try:
                    #print(target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'])               
                    if target['AbortIncompleteMultipartUpload']['DaysAfterInitiation'] > 7 or target['ID'] not in ['MMSDeletionStandardPolicy','MMSDeleteMarkers','AbortIncompleteMultipartUploadsRule','MMSVersioningPolicy']:
                        print('1-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        Rules.remove(target)
                        s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                except  KeyError:
                    continue
        elif lifecycle_config['Rules'][0]['ID'] == 'MMSVersioningPolicy':
            for target in Rules:
                try:
                    
                    if target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'] > 1 or target['NoncurrentVersionExpiration']['NoncurrentDays'] > 31 or target['Expiration']['Days'] > 31 or target['ID'] not in ['MMSDeletionStandardPolicy','MMSDeleteMarkers','AbortIncompleteMultipartUploadsRule','MMSVersioningPolicy']:
                        #print(target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'] ,target['NoncurrentVersionExpiration']['NoncurrentDays'] , target['Expiration']['Days'] > 31 or target['ID'])               
                        print('2-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        Rules.remove(target)
                        s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                except  KeyError:
                    continue
        elif lifecycle_config['Rules'][0]['ID'] == 'MMSDeletionStandardPolicy':
            for target in Rules:
                try:
                    #print(target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'])               
                    if target['Expiration']['Days'] > 2555 or target['ID'] not in ['MMSDeletionStandardPolicy','MMSDeleteMarkers','AbortIncompleteMultipartUploadsRule','MMSVersioningPolicy']:
                        print('3-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        Rules.remove(target)
                        s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                except  KeyError:
                    continue        
        
        
        policy = lifecycle_config
        Rules.append(policy['Rules'][0])
        #print(Rules)
        lcp = s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
    except ClientError as err:
        #print(err.response)
        if err.response['Error']['Code'] == 'NoSuchLifecycleConfiguration':
            s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules': lifecycle_config['Rules'] })            
        elif err.response['Error']['ArgumentName'] == 'ID':
            print( err.response['Error']['ArgumentValue'] + ' already exsists and skipping rule')


def put_bucket_lifecycle_configuration_custom(Name, lifecycle_config):
    ownerAccountId = getAccountID()
    try:
        result = s3.get_bucket_lifecycle_configuration(Bucket=Name, ExpectedBucketOwner=ownerAccountId)
        #print(result)
        Rules= result['Rules']
        if lifecycle_config['Rules'][0]['ID'] == 'MMSMoveCustomPolicy':            
            for target in Rules:
                try:
                    #print(target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'])               
                    if target['ID'] not in ['MMSDeletionStandardPolicy','MMSDeleteMarkers','AbortIncompleteMultipartUploadsRule','MMSVersioningPolicy']:
                        if target['Transitions']['StorageClass'] not in ['GLACIER_IR'] and target['Filter']['ObjectSizeGreaterThan']  > 131072  :
                            print('C1-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                            Rules.remove(target)
                            s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                except  KeyError:
                    continue
        
        policy = lifecycle_config
        Rules.append(policy['Rules'][0])
        #print(Rules)
        lcp = s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
    except ClientError as err:
        #print(err.response)
        if err.response['Error']['Code'] == 'NoSuchLifecycleConfiguration':
            s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules': lifecycle_config['Rules'] })            
        elif err.response['Error']['ArgumentName'] == 'ID':
            print( err.response['Error']['ArgumentValue'] + ' already exsists and skipping rule')



def createignorelist():
    ownerAccountId=getAccountID()
    for s1 in servicelist:
        for r1 in region:
            name='maximus-' + s1.lower() + '-logs-' + ownerAccountId + '-' + r1
            ignorelist.append(name)
            name1='maximus-' + s1.lower() + '-backup-' + ownerAccountId + '-' + r1
            ignorelist.append(name1)
            
    
    

# method to determine accountID 
def getAccountID():
    account_id = client.get_caller_identity()
    return(account_id['Account'])
    #account_id = iam.CurrentUser().arn.split(':')[4]
    #return(account_id)

def main():
    createignorelist()
    #listBuckets()
    #createXls(policy,'prechange')
    updateBucketsLcpStd()
    #listBuckets()
    #createXls(policy,'postchange')
    

## Global list variable to keep track of the Bucket Name, Transition Days, StorageClass, Status  
TransitionStatus = []

policy = {}
ruledict= {}
def getLCP(Name):
    ownerAccountId = getAccountID()
    try:
        result = s3.get_bucket_lifecycle_configuration(Bucket=Name, ExpectedBucketOwner=ownerAccountId)
        Rules = result['Rules']
        n = 1
        for r1 in Rules:
            vname= Name + '_Rule_' + str(n)
            policy[vname] = r1
            n += 1
        #print(Rules.type())            
        #policy[Name] = Rules
        #print(policy)
    except ClientError as err:
        #print(err.response['Error']['Code'])
        if err.response['Error']['Code'] == 'NoSuchLifecycleConfiguration':
            policy[Name] = {}
        
# listBuckets - lists the buckets per region and check each bucket policy using createOrUpdateLCP() method  
def listBuckets():
    BucketName = s3.list_buckets()
    #print(ignorelist)
    #for bucket in BucketName['Buckets']:
    for bucket in tqdm(BucketName['Buckets']):
        if  bucket['Name'] not in ignorelist:
            Name = bucket['Name']
            getLCP(Name)

def updateBucketsLcpStd():
    BucketName = s3.list_buckets()
    #print(ignorelist)
    #for bucket in BucketName['Buckets']:
    for bucket in tqdm(BucketName['Buckets']):
        if  bucket['Name'] not in ignorelist and bucket['Name'] == 'shankar-app-dev-s3':
            Name = bucket['Name']
            #print(MMSVersioningPolicy['Rules'][0]['ID'])
            put_bucket_lifecycle_configuration_standard(Name,MMSVersioningPolicy)
            put_bucket_lifecycle_configuration_standard(Name,MMSMPUPolicy)
            put_bucket_lifecycle_configuration_standard(Name,MMSDeleteMarkers)
            put_bucket_lifecycle_configuration_standard(Name,MMSDeletionStandardPolicy)
            
            put_bucket_lifecycle_configuration_custom(Name,MMSMoveCustomPolicy)
            

def createXls(user_dict,stage):
    currenttime = now.strftime("%H%M%S")
    filename = stage + getAccountID()+".xlsx" #+"-"+currenttime+".xlsx"
    print("Results are available in ./" + filename + ".")
    df = pd.DataFrame.from_dict(user_dict, orient='columns').transpose()
    # df = pd.DataFrame.from_dict({(i,j): user_dict[i][j] 
    #                        for i in user_dict.keys() 
    #                        for j in user_dict[i].keys()},
    #                    orient='index').transpose()
    df.to_excel(filename, index = True)
    convertxls(filename)

def convertxls(filename):
    wb = load_workbook(filename)
    ws = wb.active

    max_row=ws.max_row
    max_col=ws.max_column

    #print(max_row,max_col)
    for c1 in column_list:
        if c1 in (COL[0].value for COL in ws.iter_cols(1, ws.max_column)):
            #print(c1,COL[0].value)
            continue
        else:
            new_column = ws.max_column + 1
            #print(f'Missing Column: {c1}')
            ws.cell(row=1, column=new_column , value=c1)
            myRow = ws.row_dimensions[1]
            myRow.font = Font(bold=True)
            
    for COL in ws.iter_cols(1, ws.max_column):
        if COL[0].value == 'ID':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('this is name of the lifecycle Rule','Automation Team')
        if COL[0].value == 'AbortIncompleteMultipartUpload':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('''{'DaysAfterInitiation': 7}''','Automation Team')

        if COL[0].value == 'Filter':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('''{'ObjectSizeGreaterThan': 131072}''','Automation Team')

        if COL[0].value == 'Transitions':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('''[{'Days': 120, 'StorageClass': 'GLACIER_IR'}]''','Automation Team')
                
        if COL[0].value == 'Expiration':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('''{'Days': 121}''','Automation Team')
        if COL[0].value == 'NoncurrentVersionExpiration':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('''{'NoncurrentDays': 2555}''','Automation Team')
        if COL[0].value == 'NoncurrentVersionTransitions':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('''[{'NoncurrentDays': 31, 'StorageClass': 'STANDARD_IA', 'NewerNoncurrentVersions': 31}]''','Automation Team')

    for row in ws['A']:
        if row.value is not None :
            oldvalue=row.value
            newvalue=oldvalue.split('_')[0]
            row.value=newvalue
        
        
    wb.save(filename.split('.')[0] + 'modified.xlsx')  
if __name__ == "__main__":
    main()