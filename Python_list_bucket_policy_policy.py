import boto3,json
import logging
from botocore.exceptions import ClientError
from configparser import ConfigParser
from datetime import datetime, timedelta
import re
from openpyxl import Workbook
import pandas as pd
import json

file = "config.ini"    # give the path to the file
# config = ConfigParser()
# config.read(file)
minsize=1024000
minage=180
logging.basicConfig(level=logging.INFO,
                    format='%(levelname)s: %(asctime)s: %(message)s')
# boto3_session = boto3.session.Session(profile_name='967655172285_ie_dev_AdministratorAccess')
boto3_session = boto3.session.Session()
s3_client = boto3_session.client('s3')
s3_resource = boto3_session.resource('s3')


def visit_buckets():
    lifecycle_config_settings_it = {
        'Rules': [
            {'ID': 'S3 Intelligent Tier Transition Rule',
             'Filter': {'Prefix': ''},
             'Status': 'Enabled',
             'Transitions': [
                 {'Days': 0,
                  'StorageClass': 'INTELLIGENT_TIERING'}
             ]}
        ]}

    archive_policy = {
        'Id': 'Archive_Tier',
        'Status': 'Enabled',
        'Tierings': [
            {
                'Days': 90,
                'AccessTier': 'ARCHIVE_ACCESS'
            },
            {
                'Days': 180,
                'AccessTier': 'DEEP_ARCHIVE_ACCESS'
            }
        ]
    }
    ID = ('Archive_Tier')
    bucket_tag_key = "storage.class"
    bucket_tag_value = "s3.it"
    for bucket in s3_client.buckets.all():
        try:
            logging.info(f'bucket name = {bucket.name}')
            tag_set = s3_client.BucketTagging(bucket.name).tag_set
            for tag in tag_set:
                tag_values = list(tag.values())
                logging.info(f'tag key ={tag["Key"]}')
                logging.info(f'tag value ={tag["Value"]}')
                if (tag['Key'] == bucket_tag_key):
                    if (tag['Value'] == bucket_tag_value):
                        logging.info(f'TAG Match: tag key={bucket_tag_key} and tag value={bucket_tag_value} for bucket {bucket.name}')
                        put_bucket_lifecycle_configuration(
                            bucket.name, lifecycle_config_settings_it)
                    #   put_bucket_intelligent_tiering_configuration(bucket.name, archive_policy, ID)
        except ClientError as e:
            logging.info(f'No Tags')


def put_bucket_lifecycle_configuration(bucket_name, lifecycle_config):
    """Set the lifecycle configuration of an Amazon S3 bucket
    :param bucket_name: string
    :param lifecycle_config: dict of lifecycle configuration settings
    :return: True if lifecycle configuration was set, otherwise False
    """
    try:
        ok = boto3_session.client('s3').put_bucket_lifecycle_configuration(Bucket=bucket_name,
                                                                           LifecycleConfiguration=lifecycle_config)
        if ok:
            logging.info(f'The lifecycle configuration was set for {bucket_name}')
    except ClientError as e:
        logging.error(e)
        return False
    return True


def put_bucket_intelligent_tiering_configuration(bucket_name, archive_policy, id):
    """Set the archive tier policy for an Amazon S3 bucket
    :param bucket_name: string
    :param intel_config: dict of intelligence tier configuration settings
    :param Id = Archive ID
    :return: True if archive policy configuration was set, otherwise False
    """
    try:
        ok = boto3_session.client('s3').put_bucket_intelligent_tiering_configuration(Bucket=bucket_name,
                                                                                     IntelligentTieringConfiguration=archive_policy, Id=id)
        if ok:
            logging.info(f'The archive configuration was set for {bucket_name}')
    except ClientError as e:
        logging.error(e)
        return False
    return True


def set_object_tags(bucket, object, update=True, **new_tags):
    old_tags = {}

    if update:
        try:
            old = s3_client.get_object_tagging(Bucket=bucket, key=object)
            old_tags = {i['Key']: i['Value'] for i in old['TagSet']}
        except Exception as e:
            print(e)
            print("There was no tag")

    new_tags = {**old_tags, **new_tags}

    response = s3_client.put_object_tagging(
        Bucket=bucket,
        Key=object,
        Tagging={
            'TagSet': [{'Key': str(k), 'Value': str(v)} for k, v in new_tags.items()]
        }
    )

    print(response)


def delete_bucket_objects(Name, type, age, FileSize):
    try:
        my_bucket = s3_resource.Bucket(Name)
        for my_bucket_object in my_bucket.objects.all():
            object = s3_resource.ObjectSummary(Name, my_bucket_object.key)

            if object.size > FileSize and object.key[-3:] == type and object.last_modified.replace(tzinfo=None).isoformat() < age and object.size > minsize and object.last_modified.replace(tzinfo=None).isoformat() < minage:
                # print(object.last_modified - timedelta(days=age) )
                print(object.key, object.size,
                      object.last_modified, object.key[-3:])
                logging.info(f'Bucket = {Name}:Deleing object {object.key} from {Name} ')
                object.delete()
    except ClientError as err:
        print(err.response['Error']['Code'])


def keep_bucket_objects(Name, StorageClass, type):
    try:
        my_bucket = s3_resource.Bucket(Name)
        for my_bucket_object in my_bucket.objects.all():
            # print(my_bucket_object)
            object = s3_resource.ObjectSummary(Name, my_bucket_object.key)
            # print(object.storage_class)
            # if object.storage_class != 'INTELLIGENT_TIERING' :
            if object.key[-3:] == type:
                print(object.key, object.storage_class)
                logging.info(f'Bucket = {Name}: Adding Tag to keep object {my_bucket_object.key}')
                Tagging = {
                    'TagSet': [
                        {
                            'Key': 'Retention',
                            'Value': 'Keep',
                        },
                    ],
                }
                set_object_tags(my_bucket, my_bucket_object, Tagging)
                object.put(StorageClass=StorageClass)
                # accepted values are 'STANDARD' |'REDUCED_REDUNDANCY'|'STANDARD_IA'|'ONEZONE_IA'|'INTELLIGENT_TIERING'|'GLACIER'
    except ClientError as err:
        print(err.response['Error']['Code'])


def modify_bucket_objects(Name, StorageClass, FileSize, type):
    try:
        my_bucket = s3_resource.Bucket(Name)
        for my_bucket_object in my_bucket.objects.all():
            # print(my_bucket_object)
            object = s3_resource.ObjectSummary(Name, my_bucket_object.key)
            # print(object.storage_class)
            # if object.storage_class != 'INTELLIGENT_TIERING' :
            if object.storage_class is None and object.size >= FileSize and object.key[-3:] != type:
                print(object.key, object.storage_class, object.size)
                logging.info(f'Bucket = {Name} : Changing Storage Class of {object.key} from  STANDARD to {StorageClass}')
                object.put(StorageClass=StorageClass)
                # accepted values are 'STANDARD' |'REDUCED_REDUNDANCY'|'STANDARD_IA'|'ONEZONE_IA'|'INTELLIGENT_TIERING'|'GLACIER'
    except ClientError as err:
        print(err.response['Error']['Code'])

def is_like(text,pattern):
    regex =re.compile(pattern)
    if '%' in pattern:
        pattern = pattern.replace('%','.*?')
    if re.search(regex,text):
        return True
    return False

def populate_data(config):
    
    for item in config.sections():
        ignoreitem = config[item]['ignorebucket'].split(',')
        
    for bucket in s3_client.buckets.all():   
        #for itemcheck in ignoreitem:
            #print(f'{bucket.name}={itemcheck}={is_like(bucket.name,itemcheck)}')              
            if bucket.name not in config and bucket.name not in ignoreitem and  (re.search(bucket.name,x).group() for x in ignoreitem) :
                #if (is_like(bucket.name,x) for x in ignoreitem ):
                        logging.info(f'Adding Bucket = {bucket.name} in Config.ini file')
                        config[bucket.name] = {}
                        with open(file, 'w') as configfile:
                            config.write(configfile)


if __name__ == '__main__':
    logging.info(f'{__file__}')
    wb = Workbook()
    ws = wb.active
    r = 1
    c = 1
    
    for i in range(0,1000):      
        for j in range(0,1000):
            print (i,j,r,c)
            for bucket in s3_resource.buckets.all():
                try:
                    print('=======================')
                    print(bucket.name)
                    r1= i + r
                    c1= j+ c
                    ws.cell(row=r1,column=c1).value = bucket.name
                    print('=======================')
                    bktpolicy= s3_client.get_bucket_lifecycle(Bucket=bucket.name)
                    #json.loads(bktpolicy)
                    for dict in bktpolicy['Rules']:
                        #print(dict)
                        for key,val in dict.items():
                            c = c + 1
                            c2 = c1 + c
                            ws.cell(row=r1,column=c2).value = val
                            
                            #print(key,val)

                except ClientError as err:
                    print(err.response['Error']['Code'])
                
    wb.save(filename = "bucket_lsRules.xlsx")
    
    for bucket in s3_resource.buckets.all():
        try:
            bktpolicy= s3_client.get_bucket_lifecycle(Bucket=bucket.name)
            print(bktpolicy)
        except ClientError as err:
            print(err.response['Error']['Code'])
            
            
