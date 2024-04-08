from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
wb = load_workbook('813408048622.xlsx')
ws = wb.active

max_row=ws.max_row
max_col=ws.max_column

print(max_row,max_col)
column_list=['test2',None, 'ID', 'Filter', 'Status', 'AbortIncompleteMultipartUpload', 'Prefix', 'Expiration', 'Transitions', 'NoncurrentVersionExpiration', 'NoncurrentVersionTransitions']

for c1 in column_list:
    #for COL in ws.iter_cols(1, ws.max_column):
        if c1 in (COL[0].value for COL in ws.iter_cols(1, ws.max_column)):
            #print(c1,COL[0].value)
            continue
        else:
            new_column = ws.max_column + 1
            print(f'Missing Column: {c1}')
            ws.cell(row=1, column=new_column , value=c1)
            myRow = ws.row_dimensions[1]
            myRow.font = Font(bold=True)



for COL in ws.iter_cols(1, ws.max_column):
    if COL[0].value == 'ID':
        for i, cell in enumerate(COL):
            if i == 0:
                continue
            cell.comment = Comment('this is name of the lifecycle Rule','Automation Team')
    if COL[0].value == 'AbortIncompleteMultipartUpload':
        for i, cell in enumerate(COL):
            if i == 0:
                continue
            cell.comment = Comment('''{'DaysAfterInitiation': 7}''','Automation Team')

    if COL[0].value == 'Filter':
        for i, cell in enumerate(COL):
            if i == 0:
                continue
            cell.comment = Comment('''{'ObjectSizeGreaterThan': 131072}''','Automation Team')

    if COL[0].value == 'Transitions':
        for i, cell in enumerate(COL):
            if i == 0:
                continue
            cell.comment = Comment('''[{'Days': 120, 'StorageClass': 'GLACIER_IR'}]''','Automation Team')
            
    if COL[0].value == 'Expiration':
        for i, cell in enumerate(COL):
            if i == 0:
                continue
            cell.comment = Comment('''{'Days': 121}''','Automation Team')
    if COL[0].value == 'NoncurrentVersionExpiration':
        for i, cell in enumerate(COL):
            if i == 0:
                continue
            cell.comment = Comment('''{'NoncurrentDays': 2555}''','Automation Team')
    if COL[0].value == 'NoncurrentVersionTransitions':
        for i, cell in enumerate(COL):
            if i == 0:
                continue
            cell.comment = Comment('''[{'NoncurrentDays': 31, 'StorageClass': 'STANDARD_IA', 'NewerNoncurrentVersions': 31}]''','Automation Team')

for row in ws['A']:
    if row.value is not None :
        oldvalue=row.value
        newvalue=oldvalue.split('_')[0]
        row.value=newvalue
    
    
wb.save('813408048622.xlsx')  # + 'modified.xlsx')