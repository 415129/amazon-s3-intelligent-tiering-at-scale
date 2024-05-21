import boto3
import logging
import pandas as pd
from datetime import datetime
from botocore.exceptions import ClientError
from time import strftime 
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from  boto_formatter.core_formatter import boto_response_formatter
logname = datetime.now().strftime('logfile_S3_lcp_create_%H_%M_%S_%d_%m_%Y.log')
print(logname)
logging.basicConfig(filename=logname,level=logging.INFO,filemode='w', format='%(levelname)s: %(asctime)s: %(message)s' ,force=True)
logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client('s3')
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
iam = boto3.resource('iam')
client = boto3.client('sts')
#SGW = boto3.client('storagegateway') #,region_name='us-west-2')
ignorelist=[]
servicelist=['vpc','s3','elb','CloudTrail','rds']
region=['us-east-1','us-west-1','ca-central-1','eu-west-2','us-west-2','us-east-2','ap-south-1']
column_list=[None, 'ID', 'Filter', 'Status', 'AbortIncompleteMultipartUpload', 'Prefix', 'Expiration', 'Transitions', 'NoncurrentVersionExpiration', 'NoncurrentVersionTransitions']
stdpname = ['MMSStdVerPolicy_0D_3vR','MMSDeletionStandardPolicy','MMSStdDelMarkerPolicy','AbortIncompleteMultipartUploadsRule','MMSStdVerPolicy_31D_1vR']
## Global list variable to keep track of the Bucket Name, Transition Days, StorageClass, Status  
TransitionStatus = []

policy = {}
ruledict= {}
sgwlist=[]

MMSStdVerPolicy_31D_1vR = {
    'Rules': [
        {'ID': 'MMSStdVerPolicy_31D_1vR',
         'Expiration': {'Days': 31},
         'Filter': {},
         'Status': 'Enabled', 
         'NoncurrentVersionExpiration': {'NoncurrentDays': 31, 'NewerNoncurrentVersions': 1}
        }
    ]}

MMSStdVerPolicy_0D_3vR = {
    'Rules': [
        {'ID': 'MMSStdVerPolicy_0D_3vR',
         'Filter': {},
         'Status': 'Enabled', 
         'NoncurrentVersionExpiration': {'NoncurrentDays': 1, 'NewerNoncurrentVersions': 3}
        }
    ]}

AbortIncompleteMultipartUploadsRule = {
    'Rules': [
        {
          'ID': 'AbortIncompleteMultipartUploadsRule',
          'Expiration': {'ExpiredObjectDeleteMarker': False},
          'Filter': {},
          'Status': 'Enabled',
          'AbortIncompleteMultipartUpload': {'DaysAfterInitiation': 7}           
        }        
        ]
    }

MMSStdDelMarkerPolicy= {
    'Rules': [
        {
          'ID': 'MMSStdDelMarkerPolicy',
          'Expiration': {'ExpiredObjectDeleteMarker': True},
          'Filter': {},
          'Status': 'Enabled'         
        }        
        ]
    }

# MMSDeletionStandardPolicy = {
#     'Rules': [
#         {
#         'ID': 'MMSDeletionStandardPolicy',
#         'Expiration': {'Days': 2555},
#         'Filter': {}, 
#         'Status': 'Enabled'
#         }
#         ]
#     }

MMSStdMovPolicy_128kb_120D_G_IR_7Y  = {
    'Rules': [
        {
        'ID': 'MMSStdMovPolicy_128kb_120D_G_IR_7Y ',
        'Expiration': {'Days': 2555},
        'Filter': {'ObjectSizeGreaterThan': 131072}, 
        'Status': 'Enabled',
        'Transitions': [{'Days': 121, 'StorageClass': 'GLACIER_IR'}]
        }
        ]
    }

def put_bucket_lifecycle_configuration_standard(Name, lifecycle_config):
    ownerAccountId = getAccountID()
    try:
        result = s3.get_bucket_lifecycle_configuration(Bucket=Name, ExpectedBucketOwner=ownerAccountId)
        #print(result)
        Rules= result['Rules']
        if lifecycle_config['Rules'][0]['ID'] == 'MMSStdDelMarkerPolicy':            
            for target in Rules:
                try:
                    #print(target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'])               
                    if target['Expiration']['ExpiredObjectDeleteMarker'] == 'Ture' or target['ID'] not in stdpname:
                        print('0-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        logging.info('0-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        Rules.remove(target)
                        s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                except  KeyError:
                    continue
        elif lifecycle_config['Rules'][0]['ID'] == 'AbortIncompleteMultipartUploadsRule':
            for target in Rules:
                try:
                    #print(target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'])               
                    if target['AbortIncompleteMultipartUpload']['DaysAfterInitiation'] > 6 or target['ID'] not in stdpname :
                        print('1-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        logging.info('1-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        Rules.remove(target)
                        s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                    
                        
                except  ClientError as err:
                    print(err)
                    logging.error(f'ERROR = {err}')
                    if err.response['Error']['Code'] == 'InvalidRequest':
                        Rules.remove(target)
                        s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                except  KeyError:
                    continue
        elif lifecycle_config['Rules'][0]['ID'] == 'MMSStdVerPolicy_31D_1vR':
            for target in Rules:
                try:
                    
                    if (target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'] > 1 and target['NoncurrentVersionExpiration']['NoncurrentDays'] > 31) or target['Expiration']['Days'] > 31 or target['ID'] not in stdpname:
                        #print(target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'] ,target['NoncurrentVersionExpiration']['NoncurrentDays'] , target['Expiration']['Days'] > 31 or target['ID'])               
                        print('2-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        logging.info('2-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        Rules.remove(target)
                        s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                except  KeyError:
                    continue
        elif lifecycle_config['Rules'][0]['ID'] == 'MMSStdVerPolicy_0D_3vR':
            for target in Rules:
                try:                    
                    if (target['NoncurrentVersionExpiration']['NoncurrentDaystarget'] == 1 and target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'] > 3) or target['ID'] not in stdpname :
                        #print(target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'] ,target['NoncurrentVersionExpiration']['NoncurrentDays'] , target['Expiration']['Days'] > 31 or target['ID'])               
                        print('21-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        logging.info('21-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                        Rules.remove(target)
                        s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                    else:
                        Rules.append(lifecycle_config['Rules'][0])
                except  KeyError:
                    continue                
     
        
        
        lpolicy = lifecycle_config
        Rules.append(lpolicy['Rules'][0])
        #print(Rules)
        lcp = s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
        #print(lcp)
    except ClientError as err:
        if err.response['Error']['Code'] == 'AccessDenied':
            print ("This account does not own the bucket {}.".format(Name))
            logging.error("This account does not own the bucket {}.".format(Name))
    except ClientError as err:
        print(err.response)
        logging.error(f'EROOR = {err.response}')
        if err.response['Error']['Code'] == 'NoSuchLifecycleConfiguration':
            s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules': lifecycle_config['Rules'] })            
        elif  err.response['Error']['Code'] == 'InvalidRequest':
            print(Rules,err.response)
            logging.error(f'EROOR = {Rules} - {err.response}')
        elif err.response['Error']['ArgumentName'] == 'ID':
            print( err.response['Error']['ArgumentValue'] + ' already exists and skipping rule')
            logging.error(err.response['Error']['ArgumentValue'] + ' already exists and skipping rule')

            


def put_bucket_lifecycle_configuration_custom(Name, lifecycle_config):
    ownerAccountId = getAccountID()
    try:
        result = s3.get_bucket_lifecycle_configuration(Bucket=Name, ExpectedBucketOwner=ownerAccountId)
        #print(result)
        Rules= result['Rules']
        if lifecycle_config['Rules'][0]['ID'] == 'MMSStdMovPolicy_128kb_120D_G_IR_7Y':            
            for target in Rules:
                try:
                    #print(target['NoncurrentVersionExpiration']['NewerNoncurrentVersions'])               
                    if target['ID'] not in ['MMSDeletionStandardPolicy','MMSStdDelMarkerPolicy','AbortIncompleteMultipartUploadsRule','MMSStdVerPolicy_31D_1vR']:
                        if target['Transitions']['StorageClass'] not in ['GLACIER_IR'] and target['Filter']['ObjectSizeGreaterThan']  > 131072  :
                            print('C1-Deleteing LCP from Bucket = ' + Name + ' ,LCP = ' + target['ID'])
                            Rules.remove(target)
                            s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                except  KeyError:
                    continue
        
        policy = lifecycle_config
        Rules.append(policy['Rules'][0])
        #print(Rules)
        lcp = s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
    except ClientError as err:
        #print(err.response)
        logging.error(f'EROOR = {err.response}')
        if err.response['Error']['Code'] == 'NoSuchLifecycleConfiguration':
            s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules': lifecycle_config['Rules'] })            
        elif err.response['Error']['ArgumentName'] == 'ID':
            print( err.response['Error']['ArgumentValue'] + ' already exsists and skipping rule')
            logging.error(err.response['Error']['ArgumentValue'] + ' already exsists and skipping rule')
        elif err.response['Error']['Code'] == 'InvalidRequest':
            print(err.response)
            logging.error(f'EROOR = {err.response}')



def createignorelist():
    ownerAccountId=getAccountID()
    for s1 in servicelist:
        for r1 in region:
            name='maximus-' + s1.lower() + '-logs-' + ownerAccountId + '-' + r1
            ignorelist.append(name)
            name1='maximus-' + s1.lower() + '-backup-' + ownerAccountId + '-' + r1
            ignorelist.append(name1)
    name2=getAccountName() +'-terraform-remote-state'
    ignorelist.append(name2)
    BucketName = s3.list_buckets()
    for bucket in BucketName['Buckets']:
        #if bucket['Name'] in ['maximus-rds-backup-813408048622-us-east-1-baseline']:
        #print(bucket)
            try:
                tag_set = s3.get_bucket_tagging(Bucket=bucket['Name'])
                for tag in tag_set['TagSet']:
                    tag_values = list(tag.values())
                    #print(tag_values)
                    if (tag_values[0] == 'provisioner' and tag_values[1] == 'terraform') or tag_values[0] == 'group_nse':
                        #print(bucket['Name'])
                        ignorelist.append(bucket['Name'])
                    
                #print(ignorelist)
            except  KeyError:
                continue
            except ClientError as err:
                if err.response['Error']['Code'] == 'NoSuchTagSet':
                    print('')
                
            
          
    
    

# method to determine accountID 
def getAccountID():
    account_id = client.get_caller_identity()
    return(account_id['Account'])
    #account_id = iam.CurrentUser().arn.split(':')[4]
    #return(account_id)

def getAccountName():
    #account_id = s3.list_account_aliases()
    return( boto3.client('iam').list_account_aliases()['AccountAliases'][0])
    #account_id = iam.CurrentUser().arn.split(':')[4]
    #return(account_id)
    

    



def put_bucket_lifecycle_configuration(Name, lifecycle_config):
    ownerAccountId = getAccountID()
    try:
        result = s3.get_bucket_lifecycle_configuration(Bucket=Name, ExpectedBucketOwner=ownerAccountId)
        Rules= result['Rules']
        # Scenario #1 
        if any("Transitions" in keys for keys in Rules):
            for Rule in Rules:
                for key, value in Rule.items():
                    if (key == 'Transitions'):
                        Days = value[0]['Days']
                        StorageClass = value[0]['StorageClass']
                        TransitionStatus.append(Name)
                        TransitionStatus.append(Days)
                        TransitionStatus.append(StorageClass)
                        TransitionStatus.append('No changes made to S3 Lifecycle configuration')
        else:
            # Scenario #2
            policy = lifecycle_config 
            for p in policy['Rules']:
                for key, value in p.items():
                    if key =='Transitions':
                        Days = value[0]['Days']
                        StorageClass = value[0]['StorageClass']
                        TransitionStatus.append(Name)
                        TransitionStatus.append(Days)
                        TransitionStatus.append(StorageClass)
                        TransitionStatus.append('Updated the existing Lifecycle with Transition rule to S3 INT')

            Rules.append(policy['Rules'][0])
            #print(Rules)
            lcp = s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = {'Rules':Rules })
                        
    except ClientError as err:
        # Scenario #3
        # Catching a situation where bucket does not belong to an account
        print (err.response['Error']['Code'])
        if err.response['Error']['Code'] == 'AccessDenied':
                print ("This account does not own the bucket {}.".format(Name))
                logging.error("This account does not own the bucket {}.".format(Name))
                Days = 'N/A'
                StorageClass ='N/A'
                TransitionStatus.append(Name)
                TransitionStatus.append(Days)
                TransitionStatus.append(StorageClass)
                TransitionStatus.append("The Bucket "+Name+" does not belong to the account-"+ownerAccountId)
        elif err.response['Error']['Code'] == 'NoSuchLifecycleConfiguration':
                print("This bucket {} has no LifeCycle Configuration".format(Name)) 
                logging.error("This bucket {} has no LifeCycle Configuration".format(Name))
                policy = lifecycle_config 
                for p in policy['Rules']:
                    for key, value in p.items():
                        if key =='Transitions':
                            Days = value[0]['Days']
                            StorageClass = value[0]['StorageClass']
                            TransitionStatus.append(Name)
                            TransitionStatus.append(Days)
                            TransitionStatus.append(StorageClass)
                            TransitionStatus.append('Added a new S3 Lifecycle Transition Rule to S3 INT')
                lcp = s3.put_bucket_lifecycle_configuration(Bucket=Name, LifecycleConfiguration = policy)
        else:
            print (err.response['Error']['Code'])
            logging.error(err.response['Error']['Code'])
          
def getLCP(Name):
    ownerAccountId = getAccountID()
    try:
        result = s3.get_bucket_lifecycle_configuration(Bucket=Name, ExpectedBucketOwner=ownerAccountId)
        Rules = result['Rules']
        #print(Rules)
        n = 1
        for r1 in Rules:
            vname= Name + '_Rule_' + str(n)
            #policy[vname] = r1
            #n+= 1
            #print(policy)
            #print(r1['ID'])
            if r1['ID'] not in stdpname:
                policy[vname] = r1
                #print(policy)
            n += 1
        #print(Rules.type())            
        #policy[Name] = Rules
        #print(policy)
    except ClientError as err:
        #print(err.response['Error']['Code'])
        if err.response['Error']['Code'] == 'NoSuchLifecycleConfiguration':
            policy[Name] = {}
        
# listBuckets - lists the buckets per region and check each bucket policy using createOrUpdateLCP() method  
def listBuckets():
    BucketName = s3.list_buckets()
    #print(ignorelist)
    #for bucket in BucketName['Buckets']:
    for bucket in tqdm(BucketName['Buckets']):
        if  bucket['Name'] not in ignorelist: #and bucket['Name'] == 'oraclebkupbucket':
            Name = bucket['Name']
            getLCP(Name)

def updateBucketsLcpStd():
    BucketName = s3.list_buckets()
    #print(ignorelist)
    #for bucket in BucketName['Buckets']:
    for bucket in BucketName['Buckets']:
        if  bucket['Name'] not in ignorelist: # and bucket['Name'] in ['maximus-rds-backup-813408048622-us-east-1-baseline']:
        #if  bucket['Name'] not in ignorelist:
            Name = bucket['Name']
            print(f'Bucket Name = {Name}')
            logging.info(f'Bucket Name = {Name}')
            #print(MMSStdVerPolicy_31D_1vR['Rules'][0]['ID'])
            put_bucket_lifecycle_configuration(Name,MMSStdMovPolicy_128kb_120D_G_IR_7Y)
            put_bucket_lifecycle_configuration_standard(Name,MMSStdVerPolicy_31D_1vR)
            put_bucket_lifecycle_configuration_standard(Name,AbortIncompleteMultipartUploadsRule)
            put_bucket_lifecycle_configuration_standard(Name,MMSStdDelMarkerPolicy)
            put_bucket_lifecycle_configuration_standard(Name,MMSStdVerPolicy_0D_3vR)

            

def createXls(user_dict,stage):
    currenttime = now.strftime("%H%M%S")
    filename = stage + '_' + getAccountID() + ".xlsx" #+"-"+currenttime+".xlsx"
    print("Results are available in ./" + filename + ".")
    #print(user_dict)
    #print('-------------------------------------------------------------------------------------')
    # data = json.loads(user_dict)
    # formatted_data = json.dumps(data, indent=2)
    # print(formatted_data)
    df = pd.DataFrame.from_dict(user_dict, orient='columns').transpose()
    df.to_excel(filename, index = True)
    convertxls(filename)

def convertxls(filename):
    wb = load_workbook(filename)
    ws = wb.active

    max_row=ws.max_row
    max_col=ws.max_column

    #print(max_row,max_col)
    for c1 in column_list:
        if c1 in (COL[0].value for COL in ws.iter_cols(1, ws.max_column)):
            #print(c1,COL[0].value)
            continue
        else:
            new_column = ws.max_column + 1
            #print(f'Missing Column: {c1}')
            ws.cell(row=1, column=new_column , value=c1)
            myRow = ws.row_dimensions[1]
            myRow.font = Font(bold=True)
            
    for COL in ws.iter_cols(1, ws.max_column):
        if COL[0].value == 'ID':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('this is name of the lifecycle Rule','Automation Team')
        if COL[0].value == 'AbortIncompleteMultipartUpload':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('''{'DaysAfterInitiation': 7}''','Automation Team')

        if COL[0].value == 'Filter':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('''{'ObjectSizeGreaterThan': 131072}''','Automation Team')

        if COL[0].value == 'Transitions':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('''[{'Days': 120, 'StorageClass': 'GLACIER_IR'}]''','Automation Team')
                
        if COL[0].value == 'Expiration':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('''{'Days': 121}''','Automation Team')
        if COL[0].value == 'NoncurrentVersionExpiration':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('''{'NoncurrentDays': 2555}''','Automation Team')
        if COL[0].value == 'NoncurrentVersionTransitions':
            for i, cell in enumerate(COL):
                if i == 0:
                    continue
                cell.comment = Comment('''[{'NoncurrentDays': 31, 'StorageClass': 'STANDARD_IA', 'NewerNoncurrentVersions': 31}]''','Automation Team')
                
    # for row in range(2, max_row+1):
    #         for col in range(1, max_col+1):
    #             #print(ws.cell(row=row, column=col).value)
    #             if ws.cell(row=row, column=col).value not in [None, 'ID', 'Filter', 'Transitions']:
    #                 ws.delete_cols(col)
        
    for row in ws['A']:
        if row.value is not None :
            oldvalue=row.value
            newvalue=oldvalue.split('_')[0]
            row.value=newvalue
        
        
    wb.save(filename.split('.')[0] + '.xlsx')  
    
def delete_col_with_merged_ranges(sheet, idx):
    sheet.delete_cols(idx)

    for mcr in sheet.merged_cells:
        if idx < mcr.min_col:
            mcr.shift(col_shift=-1)
        elif idx <= mcr.max_col:
            mcr.shrink(right=1)


def delete_row_with_merged_ranges(sheet, idx):
    sheet.delete_rows(idx)

    for mcr in sheet.merged_cells:
        if idx < mcr.min_row:
            mcr.shift(row_shift=-1)
        elif idx <= mcr.max_row:
            mcr.shrink(bottom=1)
            
def storagegatewaylist():
    for r1 in region:
        SGW = boto3.client('storagegateway',region_name=r1)    
        response=SGW.list_gateways()
        #print(response)
        for gtw in response['Gateways']:
            #print(gtw['GatewayARN'],gtw['GatewayName'])
            fileShare= SGW.list_file_shares(GatewayARN=gtw['GatewayARN'])
            for fs1 in fileShare['FileShareInfoList']:
                if fs1['FileShareType'] == 'SMB':
                    #print(fs1['FileShareARN'])
                    smb_file_shares=SGW.describe_smb_file_shares(FileShareARNList=[fs1['FileShareARN']])
                    #print(smb_file_shares)
                    for fsd1 in smb_file_shares['SMBFileShareInfoList']:
                        #print(fsd1['LocationARN'].split(':::')[1])
                        ignorelist.append(fsd1['LocationARN'].split(':::')[1])
    #print(ignorelist)  
    logging.info(f'ignorelist = {ignorelist}')
                  
def main():
    storagegatewaylist()
    createignorelist()
    #print(ignorelist)
    logging.info(ignorelist)
    listBuckets()
    createXls(policy,'backup')
    policy.clear()
    updateBucketsLcpStd()
    listBuckets()
    createXls(policy,'postchange')
    
if __name__ == "__main__":
    main()